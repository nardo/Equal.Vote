# python program to output colorized voter registration maps.  Uses Python Image Library (PIL) and openpyxl
#
# Source image preparation was as follows:
#  District maps were downloaded from the state
#  The maps were passed through a Threshold filter in an image program to turn them black and white
#  Districts were filled in using the image program fill tool with the color set to R=District#, G=0, B=0
#  Images must be saved in a lossless format like .png in order for this program to work
#
# (c) 2013 Mark Frohnmayer
#
# This source code is released under the MIT License - http://opensource.org/licenses/MIT
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.

import sys
import colorsys
import Image
from openpyxl import load_workbook

base_image_name = sys.argv[1]

source_image = Image.open(base_image_name + ".png")
bounding_box = source_image.getbbox()
image_width = bounding_box[2]
image_height = bounding_box[3]

output_image_partisan_blend = Image.new("RGB", (image_width, image_height))
output_image_red_versus_blue = Image.new("RGB", (image_width, image_height))
output_image_all = Image.new("RGB", (image_width, image_height))

district_colors_blend = []
district_colors_red_v_blue = []
district_colors_all = []

blue_hue = 0.666666666
red_hue = 1

wb = load_workbook('Shutout Calculator.xlsx')
sheet = wb['Sheet1']

district_count = 60
first_district_row = 10
democrat_column = 3
republican_column = 4
total_column = 2

for index in range(0, district_count):
	democrat_count = sheet.cell(index + first_district_row, democrat_column)
	republican_count = sheet.cell(index + first_district_row, republican_column)
	total = sheet.cell(index + first_district_row, total_column)

	print democrat_count, republican_count, total
	partisan_blend_hue = blue_hue + (red_hue - blue_hue) * republican_count / (republican_count + democrat_count)
	rgb = colorsys.hsv_to_rgb(partisan_blend_hue, 1, 1)

	partisan_blend_pixel = (int(rgb[0] * 255), int(rgb[1] * 255), int(rgb[2] * 255))
	district_colors_blend.append(partisan_blend_pixel)
	
	all_blend_saturation = (democrat_count + republican_count) / float(total)
	all_blend_value = (democrat_count + republican_count) / float(total)
	#rgb = colorsys.hsv_to_rgb(partisan_blend_hue, all_blend_saturation, 1)
	rgb = colorsys.hsv_to_rgb(partisan_blend_hue, all_blend_saturation, all_blend_value)
	all_blend_pixel = (int(rgb[0] * 255), int(rgb[1] * 255), int(rgb[2] * 255))
	district_colors_all.append(all_blend_pixel)
	print all_blend_pixel
	
	district_colors_red_v_blue.append((255, 0, 0) if republican_count > democrat_count else (0, 0, 255))

for y in range(0, image_height):
	for x in range(0, image_width):
		source_pixel = source_image.getpixel((x, y))
		# District ID is encoded in the red channel of the source image -- so any pixel that has a non-zero red value and zero green and blue
		# values is the color of that numbered district
		if source_pixel[0] != 0 and source_pixel[1] == 0 and source_pixel[2] == 0:
			index = source_pixel[0] - 1
			if index < district_count:
				partisan_blend_pixel = district_colors_blend[index]
				all_blend_pixel = district_colors_all[index]
				red_blue_pixel = district_colors_red_v_blue[index]
			else:
				partisan_blend_pixel = (255,255,0)
				all_blend_pixel = (255,255,0)
				red_blue_pixel = (255,255,0)
		else:
			partisan_blend_pixel = source_pixel
			all_blend_pixel = source_pixel
			red_blue_pixel = source_pixel
			
		output_image_partisan_blend.putpixel((x, y), partisan_blend_pixel)
		output_image_red_versus_blue.putpixel((x, y), red_blue_pixel)
		output_image_all.putpixel((x, y), all_blend_pixel)

output_image_partisan_blend.save(base_image_name + ".purple.png")
output_image_red_versus_blue.save(base_image_name + ".rb.png")
output_image_all.save(base_image_name + ".all.png")
